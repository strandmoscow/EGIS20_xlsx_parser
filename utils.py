import pandas as pd
from sqlalchemy import create_engine
from atlassian import Confluence

from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter

# Заголовки для Excel (ЕГИС 1.0, короткие и удобочитаемые)
HEADERS_e1 = [
    "№ п/п",
    "Год разработки по ГК",
    "Источник",
    "Номер п. ТЗ",
    "Функциональность в соответствии с ТЗ",
    "Задокументированные функции",
    "Незадокументированные функции",
    "Используемые технологии",
    "Исходные данные для анализа",
    "Включены в ГК на эксплуатацию",
    "Востребованность системы",
    "Актуальность",
    "Уходит в ЕГИС 2.0",
    "Приоритет реализации",
    "Комментарии",
    "Код функции",
    "Новое наименование подсистемы"
]

# Заголовки для Excel (ЕГИС 2.0, короткие и удобочитаемые)
HEADERS_e2 = [
    "№ п/п",
    "Исходная система",
    "Год разработки по ГК",
    "Источник",
    "Номер п. ТЗ",
    "Функциональность в соответствии с ТЗ",
    "Задокументированные функции",
    "Незадокументированные функции",
    "Используемые технологии",
    "Исходные данные для анализа",
    "Включены в ГК на эксплуатацию",
    "Востребованность системы",
    "Актуальность",
    "Уходит в ЕГИС 2.0",
    "Приоритет реализации",
    "Комментарии",
    "Код функции",
    "Новая функциональность",
    "Новое наименование подсистемы"
]


def init_conn():
    # Читаем CSV с данными подключения
    creds_df = pd.read_csv('creds/db_cred.csv')
    creds = creds_df.iloc[0]

    # Формируем строку подключения
    database_url = (
        f"postgresql://{creds['DB_USER']}:{creds['DB_PASSWORD']}"
        f"@{creds['DB_HOST']}:{creds['DB_PORT']}/{creds['DB_NAME']}"
    )

    # Создаём движок SQLAlchemy
    engine = create_engine(database_url)
    # Возвращаем engine, с которым можно работать через .connect() или напрямую с pandas
    return engine


def init_conf_conn():
    # Читаем CSV с данными подключения
    creds_df = pd.read_csv('creds/conf_cred.csv')
    creds = creds_df.iloc[0]

    confluence = Confluence(
        url=creds['url'],
        username=creds['username'],
        password=creds['password']
    )

    return confluence

def create_xlsx(output_file, sys_code, sys_desc, df, egis_ver):

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Записываем данные без заголовков и индекса, начиная с 4-й строки (startrow=3)
        df.to_excel(writer, sheet_name=f'П_{sys_code}', index=False, header=False, startrow=3)

        worksheet = writer.sheets[f'П_{sys_code}']

        # Записываем заголовки в 3-ю строку
        HEADERS = HEADERS_e1 if (egis_ver == 1) else HEADERS_e2
        for col_num, header in enumerate(HEADERS, 1):
            worksheet.cell(row=3, column=col_num, value=header)

        # Заполняем верхние две строки и два столбца информацией о подсистеме
        worksheet['A1'] = 'Наименование ФП:'
        worksheet['B1'] = sys_desc
        worksheet['A2'] = 'Код ФП:'
        worksheet['B2'] = sys_code

        # Оформление для ячеек A1 и A2 (заливка и жирный шрифт)
        fill = PatternFill(start_color="B7D6F6", end_color="B7D6F6", fill_type="solid")
        bold_font = Font(bold=True)
        thick_side = Side(border_style="medium", color="000000")
        thin_side = Side(border_style="thin", color="000000")
        thick_border = Border(top=thick_side, bottom=thick_side, left=thick_side, right=thick_side)
        thin_border = Border(top=thin_side, bottom=thin_side, left=thin_side, right=thin_side)

        cells = [
            'A1', 'A2', 'A3', 'B3', 'C3', 'D3', 'E3', 'F3', 'G3', 'H3', 'I3', 'J3',
            'K3', 'L3', 'M3', 'N3', 'O3', 'P3', 'Q3', 'R3', 'S3'
        ]

        if egis_ver == 1:
            cells = cells[:-2]

        for cell in cells:
            worksheet[cell].fill = fill
            worksheet[cell].font = bold_font
            worksheet[cell].border = thick_border

        # Настройка ширины столбцов и перенос текста
        for idx, col in enumerate(HEADERS, 1):
            col_letter = get_column_letter(idx)
            max_length = max(
                df.iloc[:, idx-1].astype(str).map(len).max(),
                len(col)
            )
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[col_letter].width = adjusted_width

        for row in worksheet.iter_rows():
            for cell in row:
                align = cell.alignment
                cell.alignment = Alignment(
                    horizontal=align.horizontal,
                    vertical=align.vertical,
                    wrap_text=True
                )
                cell.border = thin_border
